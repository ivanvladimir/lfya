#!/usr/bin/env python
# -*- coding: utf-8
# ----------------------------------------------------------------------
# Controla orden de presentación
# ----------------------------------------------------------------------
# Ivan V. Meza
# 2018/IIMAS, México
# ----------------------------------------------------------------------
# consolidate_scores.py is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
# -------------------------------------------------------------------------
prog="presentacion"

# System libraries
import argparse
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook

if __name__ == "__main__":
    # Command line options
    p = argparse.ArgumentParser("Consolida scores from excel file")
    p.add_argument('EXCEL',
        help="Archivo con scores")
    p.add_argument("-v", "--verbose",
            action="store_true", dest="verbose",
            help="Modo verbose [Off]")
    p.add_argument('--version', action='version', version='%(prog)s 0.1')
    opts = p.parse_args()

    # Prepara función de verbose  -----------------------------------------
    if opts.verbose:
        def verbose(*args,**kargs):
            print(*args,**kargs)
    else:
        verbose = lambda *a: None

    wb = load_workbook(filename = opts.EXCEL)

    SCORES={}
    for sheetname in wb.sheetnames:
        sheet=wb[sheetname]
        rng=sheet['B2':'C'+str(sheet.max_row)]
        tmp={}
        for cell in rng:
            email=cell[0].value
            score=cell[1].value
            try:
                if tmp[email]>score:
                    tmp[email]=score
            except KeyError:
                tmp[email]=score
        for email,score in tmp.items():
            try:
                SCORES[email]+=score
            except KeyError:
                SCORES[email]=score

    for email,score in SCORES.items():
        print(email,score)



